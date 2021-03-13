"""
@author : Tej Prakash Sharma
Functions used to handle Excel
"""
import openpyxl


def get_row_count(file,sheetname):
    book=openpyxl.load_workbook(file)
    sheet=book.get_sheet_by_name(sheetname)
    return sheet.max_row


def get_column_count(file,sheetname):
    book = openpyxl.load_workbook(file)
    sheet = book.get_sheet_by_name(sheetname)
    return sheet.max_column


def read_data(file,sheetname,rowno,colno):
    book = openpyxl.load_workbook(file)
    sheet = book.get_sheet_by_name(sheetname)
    return sheet.cell(row=rowno,column=colno).value


def write_data(file, sheetname, rowno, colno, data):
    book = openpyxl.load_workbook(file)
    sheet = book.get_sheet_by_name(sheetname)
    sheet.cell(row=rowno, column=colno).value = data
    book.save(file)

















